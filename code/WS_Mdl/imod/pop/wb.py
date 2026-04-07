import os
from pathlib import Path

import flopy as fp
import numpy as np
import pandas as pd
from WS_Mdl.core.mdl import Mdl_N
from WS_Mdl.core.path import Pa_WS
from WS_Mdl.core.style import Sep, set_verbose, sprint

Pa_template = Pa_WS / r'code\WS_Mdl\Auxi\WB_Diff_NET_sum_Pkg_TEMPLATExlsx.xlsx'


def Diff_to_xlsx(
    MdlN: str,
    MdlN_B: str = None,
    date: str = None,
    cumulative=True,
    sum_Pkg: bool = True,
    net_only: bool = True,
    Pa_Out: str | None = None,
    verbose: bool = True,
    open_after: bool = True,
):  # 666 add option to do cumulative of just for that SP
    """
    Compares the water budget of two models (MdlN and MdlN_B) for a specific date and saves the differences to an Excel file.
    - date: YYYY-MM-DD
    - cumulative: if True, uses cumulative budget data; otherwise, uses daily budget data.
    - sum_Pkg: if True, sum rows where the first 3 index characters match (e.g. CHD1 + CHD2 = CHD).
    - net_only: if True, only keeps NET rows (i.e. drops IN OUT rows)
    - Pa_Out: Output path, if None defaults to M.Pa.PoP_Out_MdlN / f'WB/WB_Diff_{MdlN}_vs_{MdlN_B}_{date}.xlsx'
    """
    set_verbose(verbose)
    sprint(Sep)
    sprint(f'----- WB Diff - {MdlN}_vs_{MdlN_B} - {date} -----', set_time=True)

    # Load basics
    M = Mdl_N(MdlN)
    M_B = Mdl_N(MdlN_B) if MdlN_B else Mdl_N(M.B)
    start_date = pd.to_datetime(M.INI.SDATE, format='%Y%m%d')

    sprint(' -- Loading MF6 WB ... ', end='', set_time=True)
    # Load budget to dataframes. fp.utils.Mf6ListBudget returns a tuple. [0] is WB per SP. [1] is cumulative.
    i = 1 if cumulative else 0
    DF_1 = fp.utils.Mf6ListBudget(M.Pa.LST_Mdl).get_dataframes(
        start_datetime=start_date - pd.Timedelta(days=1), diff=net_only
    )[i]
    DF_2 = fp.utils.Mf6ListBudget(M_B.Pa.LST_Mdl).get_dataframes(
        start_datetime=start_date - pd.Timedelta(days=1), diff=net_only
    )[i]

    # Use latest common date if not specified
    if date is None:
        date = min(DF_1.index[-1], DF_2.index[-1]).strftime('%Y-%m-%d')

    # Combine DFs + Drop extra rows.
    S_1 = DF_1.loc[DF_1.index == date].squeeze()
    S_2 = DF_2.loc[DF_2.index == date].squeeze()
    S_1.index = S_1.index.str.upper()
    S_2.index = S_2.index.str.upper()

    DF_MF6 = pd.DataFrame(data={MdlN: S_1, MdlN_B: S_2})  # .squeeze(), MdlN_B: S_2.squeeze()})
    DF_MF6.drop(
        index=[
            'TOTAL_IN',
            'TOTAL_OUT',
            'PERCENT_DISCREPANCY',
            'IN-OUT',
            'TOTAL',
        ],
        inplace=True,
        errors='ignore',
    )

    if sum_Pkg:  # SUM packages (e.g. CHD1 + CHD2 = CHD)
        S_idx = DF_MF6.index.to_series().astype(str)
        S_suffix = np.where(
            S_idx.str.endswith('_IN'),
            '_IN',
            np.where(S_idx.str.endswith('_OUT'), '_OUT', '_NET'),
        )
        S_grp = S_idx.str[:3] + S_suffix
        DF_MF6 = DF_MF6.groupby(S_grp, sort=False).sum(min_count=1)

    if not net_only:  # Add net rows. If net_only, the MF functions add them in already.
        S_i = DF_MF6.index.to_series()
        S_i_in = S_i[S_i.str.endswith('_IN')]
        S_i_out = S_i[S_i.str.endswith('_OUT')]
        S_i_net = S_i_in.str.replace('_IN', '_NET')
        DF_MF6_NET = DF_MF6.loc[S_i_in.values].set_axis(S_i_net.values, axis=0) - DF_MF6.loc[S_i_out.values].to_numpy()
        DF_MF6 = pd.concat([DF_MF6, DF_MF6_NET], axis=0)

    # Sort index
    MF_indexes = [S_i_in, S_i_out, S_i_net] if not net_only else [DF_MF6.index.to_series()]
    sorted_i = pd.concat([i.sort_values() for i in MF_indexes]).values
    DF_MF6 = DF_MF6.loc[sorted_i]

    sprint('🟢\n -- Loading MSW WB ... ', end='', set_time=True, print_time_first=True)
    # Load WB for each MdlN
    DF_MSW = pd.DataFrame()
    for m in [M, M_B]:
        # Load MSW WB Out CSV and handle dates
        DF_MSW_m = pd.read_csv(m.Pa.MSW / 'msw/csv/tot_svat_dtgw.csv', index_col=False, skipinitialspace=True)
        DF_MSW_m['TimeStamp'] = pd.to_datetime(
            DF_MSW_m['TimeStamp'].str.replace(' 24:', ' 00:', regex=False)
        )  # Convert TimeStamp to datetime, handling '24:' as '00:'

        # Calculate Sum for the same period as GW budget and append to DF
        date_selection = (DF_MSW_m['TimeStamp'] >= start_date) & (DF_MSW_m['TimeStamp'] <= pd.to_datetime(date))
        DF_MSW[m.MdlN] = DF_MSW_m.loc[date_selection].iloc[:, 3:].sum()

    # Groom
    DF_MSW.rename(index={i: i.replace('(m3)', '') for i in DF_MSW.index}, inplace=True)
    d_MSW_Agg = {
        'decStot': [i for i in DF_MSW.index if ('decS' in i)],
        'Ps': ['Psgw', 'Pssw'],
        'Id': ['Idgw', 'Idsw', 'Iddgw', 'Iddsw'],
        'ETact': ['Esp', 'Eic', 'Epd', 'Ebs', 'Tact'],
        'qMF': ['qmodf', 'qsimcorrtot'],
    }  # , 'qsimtot']
    for k in d_MSW_Agg:
        DF_MSW.rename(index={i: k for i in d_MSW_Agg[k]}, inplace=True)
    DF_MSW = DF_MSW.groupby(level=0, sort=False).sum()
    DF_MSW.drop(index=['qsimtot', 'vcr'], inplace=True)

    sprint('🟢\n -- Merging & Saving ... ', end='', set_time=True, print_time_first=True)
    DF = pd.concat([DF_MF6, DF_MSW], axis=0)
    DF = DF.loc[(DF != 0).any(axis=1)]  # Drop 0 rows

    # Add descriptions
    d_Par = {
        'Pm': 'Precipitation (measured)',
        'Ps': 'Sprinkling Precipitation',
        'ETact': 'Evapotranspiration (actual)',
        'qrun': 'Runoff',
        'decStot': 'total storage change',
        'qMF': 'MODFLOW Stress',
        'RCH_NET': 'Recharge',
        'CHD_NET': 'Constant Head',
        'DRN_NET': 'Drainage',
        'RIV_NET': 'Rivers',
        'SFR_NET': 'Streams',
        'STO_NET': 'Storage',
        'WEL_NET': 'Wells',
    }
    DF['Parameter'] = DF.index.map(d_Par)
    MF_Par = [i for i in d_Par if '_NET' in i]
    MSW_Par = [i for i in d_Par if '_NET' not in i]

    # Calc Diff
    DF['Diff'] = DF[MdlN] - DF[MdlN_B]
    DF['Diff_%'] = DF.apply(
        lambda x: x['Diff'] / x[MdlN_B] * 100 if pd.notnull(x['Diff']) and x[MdlN_B] != 0 else np.nan, axis=1
    )
    DF[f'%_Pm_{MdlN}'] = DF[MdlN] / DF.loc['Pm', MdlN] * 100
    DF[f'%_Pm_{MdlN_B}'] = DF[MdlN_B] / DF.loc['Pm', MdlN_B] * 100

    # Final touches + save
    DF.insert(len(DF.columns) - 1, 'Parameter', DF.pop('Parameter'))  # Move Parameter to last column
    DF = DF.reindex(d_Par.keys())  # Reorder rows based on d_Par keys.
    date_ = date.replace('-', '')
    Pa = (
        (M.Pa.PoP_Out_MdlN / f'WB/WB_Diff_{MdlN}_m_{MdlN_B}_{date_}{"_cumulative" if cumulative else ""}')
        if Pa_Out is None
        else Pa_Out
    )

    Pa = Path(Pa).with_suffix('') if isinstance(Pa, str) else Pa.with_suffix('')
    Pa.parent.mkdir(exist_ok=True)

    # WB closing
    DF_closing = pd.DataFrame()
    DF_closing['MF SUM'] = DF.loc[MF_Par, [MdlN, MdlN_B]].sum()
    DF_closing['MSW SUM'] = DF.loc[MSW_Par, [MdlN, MdlN_B]].sum()
    DF_closing['MF ABS SUM'] = DF.loc[MF_Par, [MdlN, MdlN_B]].apply(lambda x: x.abs()).sum()
    DF_closing['MSW ABS SUM'] = DF.loc[MSW_Par, [MdlN, MdlN_B]].apply(lambda x: x.abs()).sum()
    DF_closing['MF SUM % error'] = DF_closing['MF SUM'] / DF_closing['MF ABS SUM'] * 100
    DF_closing['MSW SUM % error'] = DF_closing['MSW SUM'] / DF_closing['MSW ABS SUM'] * 100

    if sum_Pkg and net_only:
        WB_save_with_template(DF, Pa_template, 'Diff', Pa.with_suffix('.xlsx'))

    # Write to CSV
    # Reformatting values manually before saving, so CSV retains comma separators
    for col in DF.columns:
        if col not in ['Diff_%', 'Parameter']:
            DF[col] = DF[col].map(lambda x: f'{x:,.0f}' if pd.notna(x) else '')
        elif '%' in col:
            DF[col] = DF[col].map(lambda x: f'{x:.1f}' if pd.notna(x) else '')

    DF.to_csv(f'{Pa}.csv', index=True)
    DF_closing.T.to_csv(f'{Pa}_Closing.csv', index=True)

    sprint('🟢', print_time_first=True)
    sprint(Sep)

    if open_after:
        os.startfile(f'{Pa}.xlsx' if sum_Pkg and net_only else f'{Pa}.csv')

    return DF


def WB_save_with_template(DF, Pa_template, tab, Pa_Out):
    """Saves the provided DataFrame to an Excel file using the provided template."""
    from openpyxl import load_workbook

    # Load the template workbook and select the active worksheet
    wb = load_workbook(Pa_template)
    ws = wb[tab]

    # Write column headers in row 1, starting at column 2 (B)
    for c_idx, col_name in enumerate(DF.columns, start=2):
        ws.cell(row=1, column=c_idx, value=col_name)

    # Map DataFrame data to Excel cells.
    # Data starts at Row 2, Col A is the index.
    for r_idx, row_name in enumerate(DF.index, start=2):
        for c_idx, col_name in enumerate(DF.columns, start=2):
            # Overwrite the value, leaving the cell's style untouched
            val = DF.loc[row_name, col_name]

            # optional: handle NaN values before writing
            if pd.isna(val):
                val = None

            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val

            # Explicitly enforce number formats if the template cells were set to "General"
            if val is not None and col_name != 'Parameter':
                cell.number_format = '0.0' if '%' in col_name else '#,##0'

    # Save as your new file
    wb.save(Pa_Out)
